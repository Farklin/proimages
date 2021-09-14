from django.http import HttpResponse
from download.forms import ImageForm
from django.shortcuts import render, redirect
from .models import Image 
import openpyxl  
from transliterate import translit
from django.core.files.storage import FileSystemStorage
from openpyxl.styles import Color, PatternFill, Font, Border
import os
import requests 
import random 
import math 

def index(request):
    images = Image.objects.all() 
    print(images)
    return render(request, 'images.html', {'images': images})

def upload(request): 
    if request.method == 'POST':
        form = ImageForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            # Get the current instance object to display in the template
            img_obj = form.instance
        return redirect('images')
        #render(request, 'upload.html', {'form': form, 'img_obj': img_obj})
    else:
        form = ImageForm()
    return render(request, 'upload.html', {'form': form})

def view_image(request, image_id): 
   image = Image.objects.get(id=image_id)

   return render(request, 'image.html', {'id':image.id, 'title': image.title,'src': image.path_image,'date': image.date,   })

def delete_image(request, image_id): 
    try:
        image = Image.objects.get(id=image_id)
        image.path_image.delete(save=True) 
        image.delete() 
        return redirect('images')
    except: 
        return redirect('images')

    
def view_menu(request): 
    
    return render(request, 'menu.html')

# картинки 

def save_image_from_table(request): 
    

    return render(request, 'product/save_image.html')

def save_image(request): 
    if request.method == 'POST':
        try:

            image = request.POST.get("url_image")
            name = str(image).split('/')[-1] 
            
            if os.path.isfile('media/image/' + name):
                name = str(random.randrange(1, 20, 1)) + str(image).split('/')[-1] 


            image = requests.get(image).content
            
            out = open('media/image/'+ name, "wb")

            _export_excel(image, name)
            out.write(image)
            out.close()
            return HttpResponse (name)
        except: 
            _export_excel(image, '')
            return HttpResponse (False)


def _export_excel(name, new_name):

    try: 
        book = openpyxl.load_workbook('book.xlsx')
        sheet = book.active

        max_row = sheet.max_row
        sheet.cell(row = max_row + 1 , column = 1).value = name 
        sheet.cell(row = max_row + 1 , column = 2).value = new_name 
        book.save('book.xlsx')
        
    except:
        book = openpyxl.Workbook() 
        book.save('book.xlsx')
    

# @ end картинки 


# загрузка файлов xlsx 
def import_table(request): 

    files = [] 
    for file in os.listdir("media/"):
        if file.endswith(".xlsx"):
            files.append(file.replace('.xlsx', '').replace("'", ""))

    filename = '' 
    if request.method == 'POST':
        upload_file = request.FILES['excel']
        fs = FileSystemStorage()


        namefile = translit(upload_file.name, 'ru' , reversed=True).replace(' ', '_')
    
        fs.save('media/' +namefile , upload_file)


        filename =  translit(upload_file.name, reversed=True).replace('.xlsx', '').replace(' ', '_')
        if filename: 
            files.append(filename)
            return render(request, 'product/import_table.html', {'files': files})  
    else: 
        return render(request, 'product/import_table.html', {'files': files})




def delete_file_xlsx(request, filename): 
    os.remove('media/' + filename + '.xlsx')
    return redirect('import_table')

def import_product(request, filename):    
    file = 'media/' + filename + '.xlsx'


    if request.method == 'POST':
        mas_image = request.POST.getlist("mas_image[]")
        mas_har = request.POST.getlist("mas_har[]")
        mas_id = request.POST.getlist("mas_id[]")
        mas_name = request.POST.getlist("mas_name[]")
        mas_description = request.POST.getlist("mas_description[]")
        id_category = request.POST.get('id_category') 
        image_split = request.POST.get('image_split')

        unics = [] 
        names = [] 
        description = [] 
        

        print(mas_description)
        
        if mas_id != []: 
            
            for row in export(file, mas_id): 
                unics.append(row[mas_id[0]]) 

        if mas_name != []: 
            for row in export(file, mas_name): 
                names.append(row[mas_name[0]]) 
    
        if  mas_description != []: 
            for row in export(file, mas_description): 
                description.append(row[mas_description[0]])

        if mas_image != []: 
            book = openpyxl.Workbook() 
            sheet = book.active 
            
            count_row = 1
            print(image_split)
            if image_split == 'true': 
                #если картинки с делителем 

                for i, row in enumerate(export(file, mas_image)) : 
                    for id_column_image in mas_image:
                        if row[id_column_image] != None and unics[i] != None: 
                            try:  
                                for image_split in row[id_column_image].split(';'): 
                                    sheet.cell(row = count_row, column = 2 ).value = image_split
                                    sheet.cell(row = count_row, column = 1 ).value = unics[i]
                                    sheet.cell(row = count_row, column = 3 ).value = names[i]
                                    count_row += 1
                            except: 
                                sheet.cell(row = count_row, column = 2 ).value = row[id_column_image]
                                sheet.cell(row = count_row, column = 1 ).value = unics[i]
                                sheet.cell(row = count_row, column = 3 ).value = names[i]
                                count_row += 1

                #  конец если картинки с делителем 
            else: 

                # если картинки без делителя 

                for i, row in enumerate(export(file, mas_image)) : 
                    for id_column_image in mas_image:
                        if row[id_column_image] != None and unics[i] != None: 
                            
                            sheet.cell(row = count_row, column = 2 ).value = row[id_column_image]
                            sheet.cell(row = count_row, column = 1 ).value = unics[i]
                            sheet.cell(row = count_row, column = 3 ).value = names[i]
                            count_row += 1 

                #  конец если картинки без делителя


            book.save(filename + '_image.xlsx')


        if mas_har != []: 
            book = openpyxl.Workbook() 
            sheet = book.active 
            
            count_row = 2
            
            sheet.cell(row = 1, column = 1 ).value = "Уникальное значение"
            sheet.cell(row = 1, column = 2 ).value = "Наименование"
            sheet.cell(row = 1, column = 3 ).value = "Описание"
            sheet.cell(row = 1, column = 4 ).value = "Родительская категория"
            sheet.cell(row = 1, column = 5 ).value = "alias / uri"
            sheet.cell(row = 1, column = 6 ).value = "Все характеристики скопом"

            for row_id, row in enumerate(export(file, mas_har, 'har')) : 
                sheet.cell(row = count_row, column = 1 ).value = unics[row_id]
                sheet.cell(row = count_row, column = 2 ).value = names[row_id]
                
                try: 
                    sheet.cell(row = count_row, column = 3 ).value = description[row_id]
                except: 
                    sheet.cell(row = count_row, column = 3).value = '' 

                sheet.cell(row = count_row, column = 4 ).value = str(id_category)

                try:
                    sheet.cell(row = count_row, column = 5 ).value = translit(names[row_id] + '-' + str(row_id) + '-' +  str(unics[row_id]),'ru', reversed=True).lower().replace(" ", "-").replace("'", "").replace(",", "").replace(".", "")
                except: 
                    sheet.cell(row = count_row, column = 5 ).value = '' 
                
                har = '' 
                
                for i, id_column_har in enumerate(mas_har):
                    if row[id_column_har] != None and unics[row_id] != None: 
                        if row[id_column_har].split(' : ')[1] != '': 
                            har += row[id_column_har] + '\n'
                    sheet.cell(row = count_row, column = 6+i+1 ).value = row[id_column_har].split(' : ')[1]
                    sheet.cell(row = 1, column = 6+i+1 ).value = row[id_column_har].split(' : ')[0]

                sheet.cell(row = count_row, column = 6 ).value = har
                count_row += 1         

            book.save(filename + '_harakteristiks.xlsx')



        return render(request, 'product/table_product.html', {'filename': filename})
    else: 
        book = openpyxl.load_workbook(file)
        sheet = book.active 

        max_column = sheet.max_column 
        max_row = sheet.max_row

        mas_column = [] 
        for column in range(1, max_column+1): 
            cell = sheet.cell(column=column, row=1).value
            mas_column.append({'name': filter(cell, '[') , 'id': column })
            

        return render(request, 'product/table_product.html', {'column': mas_column, 'filename': filename})

def filter(znachenie, mask): 
    try:
        return znachenie.split(mask)[0]
    except: 
        return znachenie

def export(name_file, columns, type=""): 
    data = [] 

    book = openpyxl.load_workbook(name_file)

    sheet = book.active 

    max_row = sheet.max_row 

    for row in range(2, max_row+1):
        el = dict()
        for column in columns: 
            cell = sheet.cell(column = int(column), row = row).value
            if type == 'har': 
                name = sheet.cell(column = int(column), row = 1).value
                if name == None: 
                    name = '' 
                if cell == None:
                    cell = ''
                el[column] = str(filter(name, '[')) + ' : ' + str(cell)
            else: 
                el[column] = cell
        
        data.append(el)
        
    return data      




#загрузка сравнений файлов

def table_price(request): 
    files = [] 
    for file in os.listdir("media/"):
        if file.endswith(".xlsx"):
            files.append(file.replace('.xlsx', ''))

    namefile_new_price = ''
    namefile_old_price = '' 

    if request.method == 'POST':
        new_price = request.FILES['new_price_file']
        old_price = request.FILES['old_price_file']

        fs = FileSystemStorage()

        namefile_new_price  = translit(new_price.name, 'ru' , reversed=True).replace(' ', '_')
        namefile_old_price = translit(old_price.name, 'ru' , reversed=True).replace(' ', '_')

        fs.save('media/price/' +namefile_new_price , new_price)
        fs.save('media/price/' +namefile_old_price , old_price)

        context = {
            'new_price': analiz_column_table('media/price/' +namefile_new_price), 
            'file_new': 'media/price/' +namefile_new_price, 
            'old_price': analiz_column_table('media/price/' +namefile_old_price), 
            'file_old': 'media/price/' +namefile_old_price   
        }
        return render(request, 'price/price_and_price.html', {'context': context})  
    else: 
        return render(request, 'price/table_price.html', {'files': files})

#сравнение цен и вывод 
def price_and_price(request): 
    if request.method == 'POST':
        file_new = request.POST.get('file_new')
        file_old = request.POST.get('file_old')
        name_new = request.POST.get('name_new')
        name_old = request.POST.get('name_old')
        price_new = request.POST.get('price_new')
        price_old = request.POST.get('price_old')
        all_name_new= request.POST.get('all_name_new')
        all_name_old = request.POST.get('all_name_old')
        id = request.POST.get('id')

        mas_price = f_price(file_new, int(name_new), int(all_name_new), 30, int(price_new)) 
        result = f_price_and_price(int(id), file_old, int(name_old), int(all_name_old), 30, int(price_old), mas_price )
        return render(request,'price/modified.html', {'context': result} )
    else: 
        return HttpResponse('False')



#фильтр для исключения подстроки в строке 
def analiz_column_table(filename): 
    print(filename)
    book = openpyxl.load_workbook(filename)
    sheet = book.active 

    max_column = sheet.max_column 
    max_row = sheet.max_row

    mas_column = [] 
    for column in range(1, max_column+1): 
        cell = sheet.cell(column=column, row=1).value
        mas_column.append({'name': filter(cell, '[') , 'id': column })

    return mas_column 





#получение значений нового листа с ценами 
def f_price (name_file, name_nomer, all_name_nomer ,old_price_nomer, price_nomer): 
    my_path = name_file 
    my_wb_obj = openpyxl.load_workbook(my_path)
    my_sheet_obj = my_wb_obj.active 


    max_row = my_sheet_obj.max_row 

    
    mas = [] 
    for row in range(1, max_row+1):
        
        name_cell = my_sheet_obj.cell(row, name_nomer)
        price_cell = my_sheet_obj.cell(row, price_nomer)
        old_price_cell = my_sheet_obj.cell(row, old_price_nomer)
        all_name_cell = my_sheet_obj.cell(row, all_name_nomer)
       
        if(name_cell.value != None):
            try:
                p = float(price_cell.value)/10       
                round_price = math.ceil(p) * 10
                mas.append({'name': name_cell.value, 'all_name': all_name_cell.value, 'price':  str(round_price) , 'old_price':old_price_cell.value })
 
            except:
                mas.append({'name': name_cell.value, 'all_name': all_name_cell.value, 'price':  price_cell.value , 'old_price':old_price_cell.value })

    return mas 

#сравнение цен 
def f_price_and_price(id_nomer, name_file, name_nomer, all_name_nomer, old_price_nomer, price_nomer, mas_new):
    result = [] 

    def _fill_cell_color(cell, start_color, end_color): 
        fill = PatternFill(fill_type='solid',
                start_color=start_color,
                end_color=end_color)

        cell.fill = fill

    my_path = name_file
    my_wb_obj = openpyxl.load_workbook(my_path)
    my_sheet_obj = my_wb_obj.active 

    count_replacement_price = 0
    count_previous_price  = 0 

    max_row = my_sheet_obj.max_row 

    for row_new in mas_new: 
        
        for row in range(1, max_row+1):
            name_cell = my_sheet_obj.cell(row, name_nomer)
            price_cell = my_sheet_obj.cell(row, price_nomer)
            old_price_cell = my_sheet_obj.cell(row, old_price_nomer)
            all_name_cell = my_sheet_obj.cell(row,all_name_nomer)
            id_cell = my_sheet_obj.cell(row,id_nomer)

            if  len(str(row_new['name']))> 3 and str(row_new['name']) in str(name_cell.value): 
                if price_cell.value == row_new['price']: 

                    _fill_cell_color(price_cell,'fafa23', 'fafa23' )
                    count_previous_price += 1

                    
                    try: 
                        discrepancy = str(float(price_cell.value.replace(' ', '')) - float(row_new['price'].replace(' ', '')))
                    except: 
                        discrepancy = 'Ошибка'
                    result.append({
                        'id': id_cell.value, 
                        'name_new': row_new['name'], 
                        'name_old': name_cell.value, 
                        'all_name_new': row_new['all_name'], 
                        'all_name_old': all_name_cell.value,  
                        'price_new': row_new['price'], 
                        'price_old': price_cell.value,  
                        'discrepancy': discrepancy, 
                        'status': 'match'
                     })

                else: 
                    try: 
                        discrepancy = str(float(row_new['price'].replace(' ', '')) - float(price_cell.value.replace(' ', '')))
                    except: 
                        discrepancy = 'Ошибка'

                    result.append({
                    'id': id_cell.value,
                    'name_new': row_new['name'], 
                    'name_old': name_cell.value, 
                    'all_name_new': row_new['all_name'], 
                    'all_name_old': all_name_cell.value,  
                    'price_new': row_new['price'], 
                    'price_old': price_cell.value,  
                    'discrepancy': discrepancy, 
                    'status': 'modified'
                    })
                    
                    price_cell.value = row_new['price']

                    _fill_cell_color(price_cell, 'f07171', 'f07171')
        
                    count_replacement_price += 1

                   

    count_remaining_price = max_row - (count_replacement_price + count_previous_price)                

    print('Кол-во измененных цен: ' + str(count_replacement_price))
    print('Кол-во цен без изменений: ' + str(count_previous_price))
    print('Кол-во цен которые не затронуты: ' + str(count_remaining_price))

    my_wb_obj.save(name_file + ".xlsx") 

    return result 


            